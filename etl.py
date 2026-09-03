from datetime import datetime, timedelta
import os
import openpyxl
import pandas as pd

# --- CONFIGURACIÓN DE EMPRESAS Y RUTAS ---
CONFIG_RUTAS = {
    r'G:\.shortcut-targets-by-id\1H6coCC4GgCcvOGjxbvwnkh2A1xto5hcc\2026. Granjas\Registros_ Lotes Activos\Grupo Empresarial RRL_ Registros': (
        'GRUPO EMPRESARIAL RRL'
    ),
    r'G:\.shortcut-targets-by-id\1H6coCC4GgCcvOGjxbvwnkh2A1xto5hcc\2026. Granjas\Registros_ Lotes Activos\Agropecuaria Nueva del Oriente_ Registros Lotes Activos\PLANTILLAS 2026': (
        'AGROPECUARIA NUEVA DEL ORIENTE'
    ),
    r'G:\.shortcut-targets-by-id\1H6coCC4GgCcvOGjxbvwnkh2A1xto5hcc\2026. Granjas\Registros_ Lotes Activos\Agroavicola Chi-Hen_ Registros\PRODUCCION CHIHEN': (
        'AGROAVICOLA CHI-HEN'
    ),
}

ARCHIVO_SALIDA = os.path.join('DATA', 'REPORTE_AVITRACK_FINAL.xlsx')
FECHA_HOY = datetime.now().date()


# --- FUNCIONES AUXILIARES DE FORMATO ---
def formatear_fecha_estandar(valor):
    if pd.isna(valor) or str(valor).strip() in ('', '0'):
        return ''
    try:
        if isinstance(valor, (int, float)):
            if valor < 1000:
                return ''
            fecha_obj = (
                datetime(1899, 12, 30) + timedelta(days=valor)
            ).date()
        elif isinstance(valor, (datetime, pd.Timestamp)):
            fecha_obj = valor.date()
        else:
            limpio = str(valor).split('/')[0:3]
            fecha_obj = pd.to_datetime(
                '/'.join(limpio), errors='coerce', dayfirst=True
            ).date()

        return fecha_obj.strftime('%d/%m/%y') if fecha_obj else ''
    except Exception:
        return str(valor)


def formatear_edad_excel(valor):
    """Añade comilla simple para evitar errores de auto-reparación en Excel."""
    if pd.isna(valor) or str(valor).strip() in ('', '0'):
        return ''
    try:
        if isinstance(valor, str):
            return f"'{valor}"
        num = float(valor)
        semanas = int(num)
        dias = round((num - semanas) * 7)

        if dias == 0:
            res = str(semanas)
        elif dias >= 7:
            res = str(semanas + 1)
        else:
            res = f'{semanas} + {dias}/7'
        return f"'{res}"
    except Exception:
        return str(valor)


# --- EXTRACCIÓN DE DATOS CON COMENTARIOS DE CELDAS ---
def extraer_datos_archivo(ruta_archivo, razon_social):
    datos_archivo = []
    try:
        wb = openpyxl.load_workbook(
            ruta_archivo, data_only=True, keep_links=False
        )

        hojas_disponibles = {
            name.strip().upper(): name for name in wb.sheetnames
        }
        nombre_hoja_ini = next(
            (v for k, v in hojas_disponibles.items() if 'INF-INI' in k), None
        )
        nombre_hoja_dia = next(
            (v for k, v in hojas_disponibles.items() if 'DIA-PN' in k), None
        )

        if not nombre_hoja_ini or not nombre_hoja_dia:
            return []

        ws_ini = wb[nombre_hoja_ini]
        ws_dia = wb[nombre_hoja_dia]

        data_ini = list(ws_ini.iter_rows(values_only=True))

        # 1. Extracción de metadata maestra desde INF-INI
        def buscar_inf_ini(texto, ocurrencia=1):
            encontrados = []
            for r in range(min(len(data_ini), 60)):
                for c in range(min(len(data_ini[r]), 5)):
                    if texto.lower() in str(data_ini[r][c]).lower():
                        for offset in range(1, 4):
                            if (c + offset) < len(data_ini[r]):
                                v = data_ini[r][c + offset]
                                if pd.notna(v) and str(v).strip() != '':
                                    encontrados.append(v)
                                    break
            return (
                encontrados[ocurrencia - 1]
                if len(encontrados) >= ocurrencia
                else ''
            )

        info_maestra = {
            'Razon Social': razon_social,
            'Número de Lote :': str(buscar_inf_ini('Número de Lote')),
            'Línea de las Aves :': buscar_inf_ini('Línea de las Aves'),
            'Fecha de nacimiento :': formatear_fecha_estandar(
                buscar_inf_ini('Fecha de nacimiento')
            ),
            '# Pollitas :': buscar_inf_ini('# Pollitas'),
            'Orígen del Levante :': buscar_inf_ini('Orígen del Levante'),
            'Nombre de Granja (L) :': buscar_inf_ini('Nombre de Granja :', 1),
            'Ubicación Granja (L) :': buscar_inf_ini('Ubicación Granja :', 1),
            'Fecha corte a Producción :': formatear_fecha_estandar(
                buscar_inf_ini('Fecha corte a Producción')
            ),
            'Nombre de Granja (P) :': buscar_inf_ini('Nombre de Granja :', 2),
            'Ubicación Granja (P) :': buscar_inf_ini('Ubicación Granja :', 2),
        }

        f_tit = 6  # Fila 6 en Excel
        max_cols = ws_dia.max_column

        # 2. ENCONTRAR 'MORT' Y 'SALDO' CONSOLIDADO
        col_mort_consolidado = -1
        for col in range(max_cols, 0, -1):
            val_header = (
                str(ws_dia.cell(row=f_tit, column=col).value).strip().lower()
            )
            if 'mort' in val_header:
                col_mort_consolidado = col
                break

        if col_mort_consolidado == -1:
            return []

        col_ultimo_saldo = -1
        for col in range(max_cols, col_mort_consolidado - 1, -1):
            val_header = (
                str(ws_dia.cell(row=f_tit, column=col).value).strip().lower()
            )
            if 'saldo' in val_header:
                col_ultimo_saldo = col
                break

        if col_ultimo_saldo == -1:
            col_ultimo_saldo = max_cols

        headers_consolidado = []
        col_map = {}

        # Mapeo de columnas con soporte flexible para Ingreso Alimento
        for c in range(col_mort_consolidado, col_ultimo_saldo + 1):
            h_val = (
                str(ws_dia.cell(row=f_tit, column=c).value)
                .replace('\n', ' ')
                .strip()
            )
            if not h_val or h_val.lower() == 'nan':
                h_val = f'Columna_{c}'

            headers_consolidado.append((c, h_val))
            h_val_lower = h_val.lower()

            if 'trasl' in h_val_lower or 'ventas' in h_val_lower:
                col_map['trasl'] = c
            # BÚSQUEDA MEJORADA PARA INGRESO ALIMENTO
            elif any(p in h_val_lower for p in ['ingreso b', 'ingreso alimento', 'ingres']) and 'bandeja' not in h_val_lower:
                if 'ingreso_alim' not in col_map:
                    col_map['ingreso_alim'] = c
            elif 'salida huevos' in h_val_lower or ('salida' in h_val_lower and 'huevo' in h_val_lower):
                col_map['salida_huevo'] = c
            elif 'ingres' in h_val_lower and ('bandeja' in h_val_lower or c == col_ultimo_saldo - 3):
                col_map['ingreso_bandeja'] = c

        # 4. UBICAR FILA INICIAL
        fila_inicio = -1
        max_rows = ws_dia.max_row

        for r in range(f_tit + 1, max_rows + 1):
            f_raw = ws_dia.cell(row=r, column=2).value  # Columna B (Fecha)
            if (
                pd.notna(f_raw)
                and str(f_raw).strip() != ''
                and 'total' not in str(f_raw).lower()
            ):
                fila_inicio = r
                break

        if fila_inicio == -1:
            return []

        # 5. EXTRACCIÓN DIARIA DE DATOS Y COMENTARIOS
        for r in range(fila_inicio, max_rows + 1):
            f_raw = ws_dia.cell(row=r, column=2).value
            f_form = formatear_fecha_estandar(f_raw)

            if not f_form or 'total' in str(f_raw).lower():
                break

            try:
                fecha_dt = datetime.strptime(f_form, '%d/%m/%y').date()
                if fecha_dt > FECHA_HOY:
                    break
            except Exception:
                pass

            valores_fila = {}
            for col_idx, h_name in headers_consolidado:
                val = ws_dia.cell(row=r, column=col_idx).value
                valores_fila[h_name] = val

            # Función mejorada para obtener comentario de celda o valor en celda contigua
            def obtener_comentario_celda(col_key):
                if col_key in col_map:
                    target_col = col_map[col_key]
                    cell = ws_dia.cell(row=r, column=target_col)
                    
                    # 1. Buscar Comentario Flotante de Celda (Cell Comment)
                    if cell.comment and cell.comment.text:
                        text_comm = str(cell.comment.text).strip().replace('\n', ' ')
                        # Limpiar nombre de autor si existe "Autor:"
                        if ':' in text_comm and len(text_comm.split(':')[0]) < 20:
                            text_comm = text_comm.split(':', 1)[1].strip()
                        if text_comm:
                            return text_comm
                    
                    # 2. Si no hay comentario flotante, revisar la columna inmediatamente siguiente (+1) por si el comentario está escrito como texto
                    cell_next = ws_dia.cell(row=r, column=target_col + 1)
                    if cell_next.value and not str(cell_next.value).replace('.', '').isdigit():
                        val_str = str(cell_next.value).strip()
                        if val_str and val_str.lower() not in ['nan', '0', '0.0', 'none']:
                            return val_str
                            
                return ''

            reg = {
                **info_maestra,
                'Fecha': f_form,
                'Edad Sem + Días': formatear_edad_excel(
                    ws_dia.cell(row=r, column=3).value
                ),
            }

            for _, h_name in headers_consolidado:
                reg[h_name] = valores_fila[h_name]

            # Inserción asegurada de comentarios
            reg['Comentario_Trasl_Ventas'] = obtener_comentario_celda('trasl')
            reg['Comentario_Ingreso_Alimento'] = obtener_comentario_celda('ingreso_alim')
            reg['Comentario_Salida_Huevo'] = obtener_comentario_celda('salida_huevo')
            reg['Comentario_Entrada_Bandeja'] = obtener_comentario_celda('ingreso_bandeja')

            reg['Archivo'] = os.path.basename(ruta_archivo)
            datos_archivo.append(reg)

        wb.close()

    except Exception as e:
        print(f'⚠️ Error procesando {os.path.basename(ruta_archivo)}: {e}')

    return datos_archivo


# --- EJECUCIÓN PRINCIPAL ---
def ejecutar():
    print('--- 🚀 EXTRACCIÓN TOTAL CON COMENTARIOS DE CELDAS ---')
    os.makedirs('DATA', exist_ok=True)
    consolidado_final = []

    for ruta, nombre_empresa in CONFIG_RUTAS.items():
        if not os.path.exists(ruta):
            print(f'⚠️ Ruta no encontrada: {ruta}')
            continue

        for raiz, _, archivos in os.walk(ruta):
            for f in archivos:
                if f.endswith('.xlsx') and not f.startswith('~$'):
                    print(f'   📂 Procesando: {f}')
                    res = extraer_datos_archivo(
                        os.path.join(raiz, f), nombre_empresa
                    )
                    if res:
                        consolidado_final.extend(res)

    if consolidado_final:
        df_salida = pd.DataFrame(consolidado_final)
        df_salida.to_excel(ARCHIVO_SALIDA, index=False, engine='openpyxl')
        print(
            f'\n✅ ETL COMPLETADA. Registros extraídos: {len(df_salida)} | Archivo: {ARCHIVO_SALIDA}'
        )
    else:
        print('\n⚠️ No se encontraron datos para procesar.')


if __name__ == '__main__':
    ejecutar()